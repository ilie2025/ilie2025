import openpyxl
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timedelta
import random
import logging
import pulp

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

def citeste_date(fisier_excel):
    try:
        wb = openpyxl.load_workbook(fisier_excel)
        ws_distributie = wb["Distribuție Echitabilă"]
        ws_concedii = wb["Concedii"]

        angajati = [ws_distributie.cell(row=i, column=1).value for i in range(2, ws_distributie.max_row + 1)]
        concedii = {
            row[0]: [datetime.strptime(d, "%d-%m-%Y") for d in row[1].split(",")]
            for row in ws_concedii.iter_rows(min_row=2, values_only=True) if row[1]
        }

        return wb, angajati, concedii
    except Exception as e:
        logging.error(f"Eroare la citirea datelor din Excel: {e}")
        raise

def amesteca_angajati(angajati):
    random.shuffle(angajati)
    return angajati

def verifica_disponibilitate(angajat, zi, concedii, alocari_precedente):
    zi_precedenta = zi - timedelta(days=1)
    if alocari_precedente.get(angajat) == zi_precedenta:
        return False  # Pauză de 48h nerespectată
    if angajat in concedii and zi in concedii[angajat]:
        return False  # Este în concediu
    return True

def optimizeaza_distributia(angajati, ore_vineri, ore_sambata, ore_duminica, ture_24h):
    model = pulp.LpProblem("Echilibrare_Distributie", pulp.LpMinimize)

    ore_weekend = {
        ang: ore_vineri.get(ang, 0) + ore_sambata.get(ang, 0) + ore_duminica.get(ang, 0)
        for ang in angajati
    }
    avg_ore_weekend = sum(ore_weekend.values()) / len(angajati)

    deviations = pulp.LpVariable.dicts("Deviations", angajati, lowBound=0)

    model += pulp.lpSum([deviations[ang] for ang in angajati])

    for ang in angajati:
        model += ore_weekend[ang] + deviations[ang] >= avg_ore_weekend
        model += ore_weekend[ang] - deviations[ang] <= avg_ore_weekend

    model.solve()

    return {ang: ore_weekend[ang] for ang in angajati}

def aloca_angajati(angajati, zi, concedii, alocari_precedente, ore_weekend):
    disponibili = [
        ang for ang in angajati if verifica_disponibilitate(ang, zi, concedii, alocari_precedente)
    ]

    if len(disponibili) < 2:
        logging.warning(f"Nu sunt suficienți angajați disponibili pentru ziua {zi.strftime('%d-%m-%Y')}.")
        return []

    disponibili = amesteca_angajati(disponibili)
    disponibili.sort(key=lambda ang: (ore_weekend.get(ang, 0), alocari_precedente.get(ang, datetime(1900, 1, 1))))
    return disponibili[:2]

def actualizeaza_raport(ws_raport, angajati, ore_vineri, ore_sambata, ore_duminica, ture_24h):
    for i, ang in enumerate(angajati, start=2):
        if ws_raport.cell(row=i, column=1).value == ang:
            ws_raport.cell(row=i, column=2, value=ws_raport.cell(row=i, column=2).value + ore_vineri.get(ang, 0))
            ws_raport.cell(row=i, column=3, value=ws_raport.cell(row=i, column=3).value + ore_sambata.get(ang, 0))
            ws_raport.cell(row=i, column=4, value=ws_raport.cell(row=i, column=4).value + ore_duminica.get(ang, 0))
            ws_raport.cell(row=i, column=5, value=ws_raport.cell(row=i, column=5).value + ture_24h.get(ang, 0))
        else:
            ws_raport.cell(row=i, column=1, value=ang)
            ws_raport.cell(row=i, column=2, value=ore_vineri.get(ang, 0))
            ws_raport.cell(row=i, column=3, value=ore_sambata.get(ang, 0))
            ws_raport.cell(row=i, column=4, value=ore_duminica.get(ang, 0))
            ws_raport.cell(row=i, column=5, value=ture_24h.get(ang, 0))

def actualizeaza_diferente(ws_diferente, angajati, weekend_days):
    for i, ang in enumerate(angajati, start=2):
        ws_diferente.cell(row=i, column=1, value=ang)
        ws_diferente.cell(row=i, column=2, value=weekend_days[ang])

def genereaza_planificare_ture(fisier_excel, luna, an):
    try:
        wb, angajati, concedii = citeste_date(fisier_excel)
        ws_raport = wb["Raport Anual"] if "Raport Anual" in wb.sheetnames else wb.create_sheet("Raport Anual")
        ws_diferente = wb["Diferente Weekend"] if "Diferente Weekend" in wb.sheetnames else wb.create_sheet("Diferente Weekend")

        prima_zi = datetime(an, luna, 1)
        ultima_zi = (prima_zi + timedelta(days=31)).replace(day=1) - timedelta(days=1)
        zile_luna = [prima_zi + timedelta(days=i) for i in range((ultima_zi - prima_zi).days + 1)]

        ws_program = wb.create_sheet(title=f"Program Lunar {luna}_{an}")

        ws_program.cell(row=1, column=1, value="Nume Angajat")
        for i, zi in enumerate(zile_luna, start=2):
            ws_program.cell(row=1, column=i, value=zi.day)
        zile_saptamana = ['L', 'Ma', 'Mi', 'J', 'V', 'S', 'D']
        for i, zi in enumerate(zile_luna, start=2):
            ws_program.cell(row=2, column=i, value=zile_saptamana[zi.weekday()])

        for i, ang in enumerate(angajati, start=3):
            ws_program.cell(row=i, column=1, value=ang)

        alocari_precedente = {}
        ore_vineri, ore_sambata, ore_duminica, ture_24h = {ang: 0 for ang in angajati}, {}, {}, {}
        ore_weekend = {ang: 0 for ang in angajati}
        weekend_days = {ang: ws_diferente.cell(row=i, column=2).value or 0 for i, ang in enumerate(angajati, start=2)}

        for zi in zile_luna:
            zi_text = zi.strftime("%A")
            alocari = aloca_angajati(angajati, zi, concedii, alocari_precedente, ore_weekend)

            if len(alocari) < 2:
                continue

            for angajat in alocari:
                zi_col = zile_luna.index(zi) + 2
                ws_program.cell(row=angajati.index(angajat) + 3, column=zi_col, value="X")
                ture_24h[angajat] = ture_24h.get(angajat, 0) + 1

                if zi.weekday() == 4:  # Vineri
                    ore_vineri[angajat] = ore_vineri.get(angajat, 0) + 8
                elif zi.weekday() == 5:  # Sâmbătă
                    ore_sambata[angajat] = ore_sambata.get(angajat, 0) + 24
                elif zi.weekday() == 6:  # Duminică
                    ore_duminica[angajat] = ore_duminica.get(angajat, 0) + 16

                alocari_precedente[angajat] = zi

        ore_weekend = optimizeaza_distributia(angajati, ore_vineri, ore_sambata, ore_duminica, ture_24h)

        actualizeaza_raport(ws_raport, angajati, ore_vineri, ore_sambata, ore_duminica, ture_24h)
        actualizeaza_diferente(ws_diferente, angajati, weekend_days)

        wb.save(fisier_excel)
        logging.info(f"Planificarea pentru luna {luna}/{an} a fost generată cu succes!")

    except Exception as e:
        logging.error(f"Eroare la generarea planificării: {e}")
        raise

if __name__ == "__main__":
    FISIER_EXCEL = r"Planificare_Ture.xlsx"
    LUNA = 7
    AN = 2025
    genereaza_planificare_ture(FISIER_EXCEL, LUNA, AN)
